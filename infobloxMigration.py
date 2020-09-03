import requests
import pprint
import getpass
import wexpect
import openpyxl
import datetime
import time
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

fileName = 'Infoblox_Migration_Prep.xlsx'

def getDeviceList():
    routerList = list()
    device = requests.get('https://akips11.hsnet.ufl.edu/api-script?password=1r0nM@1d3n;function=web_export_device_list;', verify=False)

    for elem in device.text.splitlines():
        line = elem.split(',')
        temp = list()
        if line[0][1].startswith('R'):
            temp.append(line[0])
            temp.append(line[1].strip())
            routerList.append(temp)
    
    # sh ip int bri해서 Vlan 정보가 없는 경우 skip하도록 조건 설정
    #pprint.pprint(routerList)
    
    return routerList

def accessJumpBox(username, password):

    print('\n--- Attempting connection to ' + 'IS6 Server... ')
    ssh_newkey = 'Are you sure you want to continue connecting'
    session = wexpect.spawn('ssh ' + username + '@is6.hsnet.ufl.edu')

    idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

    if idx == 0:
        session.sendline('yes')
        idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

        if idx == 0:
            session.sendline(password)
    elif idx == 1:
        session.sendline(password)

    idx = session.expect(['$', wexpect.EOF])

    if idx == 0:
        print("--- Successful Login to JumpBox")
        return session
    else:
        print("--- Terminated program")
        exit()

def accessSwitches(session, switch, username, password):

    session.sendline('ssh ' + switch)

    print('\n------------------------------------------------------')
    print('--- Attempting connection to: ' + switch)
    print('------------------------------------------------------\n')

    session.expect(['word', wexpect.EOF])
    session.sendline(password)
        
    print('--- Success Login to: ', switch)
 
    idx = session.expect(['>', '#', wexpect.EOF])

    if idx == 0:
        session.sendline('en')
        idx = session.expect(['word:', wexpect.EOF])
        session.sendline(password)
        session.expect(['#', wexpect.EOF])
    
    return session

def createExcelFile():
    
    # Excel File Creation
    nowDate = 'Report Date: ' + str(datetime.datetime.now().strftime('%Y-%m-%d'))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'DHCP_Helper_Address'
    
    # Pretty display for the File
    font = Font(bold=True)
    alignment = Alignment(horizontal='center')
    bgColor = PatternFill(fgColor='BFBFBFBF', patternType='solid')
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    ws['A2'] = nowDate
    
    ws['A4'] = 'Hostname'
    ws['A4'].alignment = alignment
    ws['A4'].font = font
    ws['A4'].fill = bgColor
    ws['A4'].border = border

    ws['B4'] = 'IP Address'
    ws['B4'].alignment = alignment
    ws['B4'].font = font
    ws['B4'].fill = bgColor
    ws['B4'].border = border

    ws['C4'] = 'Vlan'
    ws['C4'].alignment = alignment
    ws['C4'].font = font  
    ws['C4'].fill = bgColor
    ws['C4'].border = border

    ws['D4'] = 'Bluecat Helper Addr'
    ws['D4'].alignment = alignment
    ws['D4'].font = font  
    ws['D4'].fill = bgColor
    ws['D4'].border = border

    ws['E4'] = 'Infoblox Helper Addr'
    ws['E4'].alignment = alignment
    ws['E4'].font = font  
    ws['E4'].fill = bgColor
    ws['E4'].border = border

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 60
    
    wb.save(fileName)
    wb.close()

def saveExcelFile(deviceList, helerAddrList, cellNumber):

    wb = openpyxl.load_workbook(fileName)
    ws = wb.active
    alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

    ws['A' + str(cellNumber)] = deviceList[0]
    ws['A' + str(cellNumber)].alignment = alignment
    ws['A' + str(cellNumber)].border = border

    ws['B' + str(cellNumber)] = deviceList[1]
    ws['B' + str(cellNumber)].alignment = alignment
    ws['B' + str(cellNumber)].border = border

    for elem in helerAddrList:
        ws['C' + str(cellNumber)] = elem[0]
        ws['C' + str(cellNumber)].alignment = alignment
        ws['C' + str(cellNumber)].border = border

        ws['D' + str(cellNumber)] = elem[1]
        ws['D' + str(cellNumber)].alignment = alignment
        ws['D' + str(cellNumber)].border = border
        
        cellNumber += 1

    wb.save(fileName)

    print('--- Data successfully saved')
    wb.close()

def getVlanList(session, switchName):

    vlanList = list()
    command = 'sh ip int bri | i up'

    session.sendline('term length 0')
    session.expect(['#', wexpect.EOF])
        
    session.sendline(command)
    session.expect(['#', wexpect.EOF])

    data = session.before.splitlines()

    if len(data[1:]) == 1:
        dump = data[1:][0].split(' ')
        for elem in dump:
            if elem.startswith('Vl'):
                vlanList.append(elem.split(' ')[0])
    else:
        for elem in data:
            if elem.startswith('Vl'):
                vlanList.append(elem.split(' ')[0])

    print("--- Complete Gathering Vlan Information")
    return vlanList

def getHelperAddr(session, vlanList):

    helperAddr = list()

    for vlan in vlanList:
        
        command = f'sh run int {vlan}'

        session.sendline(command)
        session.expect(['#', wexpect.EOF])

        for elem in session.before.splitlines()[1:]:
            if elem.strip().startswith('ip helper-address') or \
                elem.strip().startswith('ip dhcp relay address'):
                temp = list()
                temp.extend([vlan, elem])
                print(temp)
                helperAddr.append(temp)
    
    print("--- Complete Gathering Helpder Address")
    return helperAddr

if __name__ == '__main__':

    cellNumber = 5
    print()
    print('+-------------------------------------------------------------+')
    print('|    Infoblox Migration Prepartion tool...                    |')
    print('|    Helpder addres addition / ACL addition                   |')
    print('|    Version 1.0.0                                            |')
    print('|    Compatible with C35xx, C37xx, C38xx, C65XX               |')
    print('|    Nexus 5K, 7K, 9K                                         |')
    print('|    Scripted by Ethan Park, Sep. 2020                        |')
    print('+-------------------------------------------------------------+')
    print()
    username = input("Enter your admin ID ==> ")
    password = getpass.getpass("Enter your password ==> ")

    print()

    switchList = getDeviceList()
    #createExcelFile()

    for elem in switchList:
        session = accessJumpBox(username, password)
        session = accessSwitches(session, elem[1], username, password)
        vlanList = getVlanList(session, elem[0])
        helperList = getHelperAddr(session, vlanList)

        if len(helperList) != 0:
            saveExcelFile(elem, helperList, cellNumber)
            cellNumber += len(helperList)
        session.close()